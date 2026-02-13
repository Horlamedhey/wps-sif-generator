import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.models import EmployeeRow, SIFRequest
from app.sif import build_sif_rows, build_xlsx_bytes, default_filename


class SIFServiceTests(unittest.TestCase):
    def test_default_filename_format(self):
        name = default_filename("fg67", "BMCT", date(2026, 2, 13), 1)
        self.assertEqual(name, "SIF_fg67_BMCT_20260213_001.xlsx")

    def test_normalization_and_note_autofill(self):
        payload = SIFRequest(
            employer_cr="fg67",
            payer_cr="fg67",
            payer_bank_short="BMCT",
            payer_account="123",
            salary_year=2026,
            salary_month=2,
            payment_type="Salary",
            processing_date=date(2026, 2, 13),
            seq=1,
            sheet_name="Sheet1",
            employees=[
                EmployeeRow(
                    employee_id_type="x",
                    employee_name="Test User",
                    net_salary="0",
                    basic_salary="0",
                    extra_hours="1.258",
                    extra_income="10",
                    deductions="1",
                    social_security_deductions="0",
                    notes_comments="",
                )
            ],
        )

        rows, normalized, _, total_salaries, number_of_records = build_sif_rows(payload)
        self.assertEqual(number_of_records, 1)
        self.assertEqual(total_salaries, "0.000")
        self.assertEqual(normalized[0].employee_id_type, "C")
        self.assertEqual(str(normalized[0].extra_hours), "1.26")
        self.assertEqual(normalized[0].notes_comments, "Net salary is 0")
        self.assertEqual(len(rows[0]), 15)
        self.assertEqual(len(rows[1]), 15)
        self.assertEqual(len(rows[2]), 15)

    def test_workbook_single_sheet_and_headers(self):
        payload = SIFRequest(
            employer_cr="fg67",
            payer_cr="fg67",
            payer_bank_short="BMCT",
            payer_account="123",
            salary_year=2026,
            salary_month=2,
            payment_type="Salary",
            processing_date=date(2026, 2, 13),
            seq=1,
            sheet_name="Sheet1",
            employees=[EmployeeRow(employee_name="A")],
        )
        rows, _, _, _, _ = build_sif_rows(payload)
        binary = build_xlsx_bytes(rows, "Sheet1")

        wb = load_workbook(filename=BytesIO(binary))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb[wb.sheetnames[0]]
        self.assertEqual(ws.cell(row=1, column=1).value, "Employer CR-NO")
        self.assertEqual(ws.cell(row=3, column=1).value, "Employee ID Type")


if __name__ == "__main__":
    unittest.main()
